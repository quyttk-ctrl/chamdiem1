"""
Tạo hệ thống chấm điểm tự động cho bài thi
- CLO 1: Câu 1-40 (mỗi câu 0.25 điểm, tổng 10 điểm)
- CLO 2: Câu 41-55 (mỗi câu 0.5 điểm, tổng 7.5 điểm)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_scoring_system():
    # Tạo workbook mới
    wb = openpyxl.Workbook()
    
    # Xóa sheet default
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # ========== SHEET 1: CONFIG ==========
    ws_config = wb.create_sheet("Config", 0)
    
    # Header
    ws_config['A1'] = "CẤU HÌNH HỆ THỐNG CHẤM ĐIỂM"
    ws_config['A1'].font = Font(bold=True, size=14)
    ws_config.merge_cells('A1:D1')
    
    # Phần 1: Định nghĩa CLO
    ws_config['A3'] = "ĐỊNH NGHĨA CHUẨN ĐẦU RA (CLO)"
    ws_config['A3'].font = Font(bold=True, size=11)
    
    headers = ['CLO', 'Tên CLO', 'Câu từ', 'Câu đến', 'Điểm/câu', 'Tổng điểm']
    for col, header in enumerate(headers, 1):
        cell = ws_config.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Dữ liệu CLO
    clo_data = [
        ['CLO 1', 'Kiến thức cơ bản', 1, 40, 0.25, 10],
        ['CLO 2', 'Ứng dụng nâng cao', 41, 55, 0.5, 7.5]
    ]
    
    for row_idx, data in enumerate(clo_data, 5):
        for col_idx, value in enumerate(data, 1):
            ws_config.cell(row=row_idx, column=col_idx).value = value
    
    # Phần 2: Quản lý mã đề
    ws_config['A8'] = "QUẢN LÝ ĐÁP ÁN THEO MÃ ĐỀ"
    ws_config['A8'].font = Font(bold=True, size=11)
    
    headers2 = ['Mã đề', 'Câu 1', 'Câu 2', 'Câu 3', 'Câu 4', 'Câu 5']
    for col, header in enumerate(headers2, 1):
        cell = ws_config.cell(row=9, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    # Ví dụ mã đề A
    example_a = ['ĐHXD-2024-A', 'A', 'B', 'C', 'D', 'A']
    for col_idx, value in enumerate(example_a, 1):
        ws_config.cell(row=10, column=col_idx).value = value
    
    # Ví dụ mã đề B
    example_b = ['ĐHXD-2024-B', 'C', 'D', 'A', 'B', 'C']
    for col_idx, value in enumerate(example_b, 1):
        ws_config.cell(row=11, column=col_idx).value = value
    
    ws_config['A12'] = "(Tiếp tục thêm các cột cho Câu 6-55...)"
    ws_config['A12'].font = Font(italic=True)
    
    # Điều chỉnh độ rộng cột
    ws_config.column_dimensions['A'].width = 15
    for col in range(2, 7):
        ws_config.column_dimensions[get_column_letter(col)].width = 10
    
    # ========== SHEET 2: CHẤM ĐIỂM ==========
    ws_score = wb.create_sheet("Chấm Điểm", 1)
    
    # Header
    ws_score['A1'] = "BẢNG CHẤM ĐIỂM BÀI THI"
    ws_score['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_score['A1'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    ws_score.merge_cells('A1:F1')
    
    # Thông tin sinh viên
    ws_score['A3'] = "Mã sinh viên:"
    ws_score['B3'] = ""
    ws_score['A4'] = "Họ tên:"
    ws_score['B4'] = ""
    ws_score['A5'] = "Mã đề:"
    ws_score['B5'] = ""
    
    for row in [3, 4, 5]:
        ws_score.cell(row=row, column=1).font = Font(bold=True)
    
    # Bảng chấm điểm
    headers3 = ['STT', 'Câu', 'Đáp án đúng', 'Đáp án SV', 'Đúng/Sai', 'Điểm']
    for col, header in enumerate(headers3, 1):
        cell = ws_score.cell(row=7, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Tạo dòng cho 55 câu
    for i in range(1, 56):
        row = 7 + i
        ws_score.cell(row=row, column=1).value = i
        ws_score.cell(row=row, column=2).value = f"Câu {i}"
        ws_score.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws_score.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        
        # Công thức kiểm tra đúng/sai
        ws_score.cell(row=row, column=5).value = f'=IF(C{row}=D{row},"✓","✗")'
        ws_score.cell(row=row, column=5).alignment = Alignment(horizontal="center")
        
        # Công thức tính điểm
        if i <= 40:
            ws_score.cell(row=row, column=6).value = f'=IF(E{row}="✓",0.25,0)'
        else:
            ws_score.cell(row=row, column=6).value = f'=IF(E{row}="✓",0.5,0)'
        ws_score.cell(row=row, column=6).alignment = Alignment(horizontal="center")
    
    # Tổng kết
    summary_row = 7 + 56
    ws_score.cell(row=summary_row, column=5).value = "Tổng điểm:"
    ws_score.cell(row=summary_row, column=5).font = Font(bold=True, size=11)
    ws_score.cell(row=summary_row, column=5).alignment = Alignment(horizontal="right")
    
    ws_score.cell(row=summary_row, column=6).value = f'=SUM(F8:F62)'
    ws_score.cell(row=summary_row, column=6).font = Font(bold=True, size=12)
    ws_score.cell(row=summary_row, column=6).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws_score.cell(row=summary_row, column=6).alignment = Alignment(horizontal="center")
    
    # Điểm CLO
    clo1_row = summary_row + 2
    ws_score.cell(row=clo1_row, column=5).value = "Điểm CLO 1 (Câu 1-40):"
    ws_score.cell(row=clo1_row, column=5).font = Font(bold=True)
    ws_score.cell(row=clo1_row, column=5).alignment = Alignment(horizontal="right")
    
    ws_score.cell(row=clo1_row, column=6).value = f'=SUM(F8:F47)'
    ws_score.cell(row=clo1_row, column=6).font = Font(bold=True)
    ws_score.cell(row=clo1_row, column=6).fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    ws_score.cell(row=clo1_row, column=6).alignment = Alignment(horizontal="center")
    
    clo2_row = clo1_row + 1
    ws_score.cell(row=clo2_row, column=5).value = "Điểm CLO 2 (Câu 41-55):"
    ws_score.cell(row=clo2_row, column=5).font = Font(bold=True)
    ws_score.cell(row=clo2_row, column=5).alignment = Alignment(horizontal="right")
    
    ws_score.cell(row=clo2_row, column=6).value = f'=SUM(F48:F62)'
    ws_score.cell(row=clo2_row, column=6).font = Font(bold=True)
    ws_score.cell(row=clo2_row, column=6).fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    ws_score.cell(row=clo2_row, column=6).alignment = Alignment(horizontal="center")
    
    # Điều chỉnh độ rộng cột
    ws_score.column_dimensions['A'].width = 8
    ws_score.column_dimensions['B'].width = 10
    ws_score.column_dimensions['C'].width = 15
    ws_score.column_dimensions['D'].width = 15
    ws_score.column_dimensions['E'].width = 12
    ws_score.column_dimensions['F'].width = 12
    
    # ========== SHEET 3: HƯỚNG DẪN ==========
    ws_guide = wb.create_sheet("Hướng Dẫn", 2)
    
    ws_guide['A1'] = "HƯỚNG DẪN SỬ DỤNG HỆ THỐNG CHẤM ĐIỂM"
    ws_guide['A1'].font = Font(bold=True, size=12, color="FFFFFF")
    ws_guide['A1'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    ws_guide.merge_cells('A1:B1')
    
    row = 3
    guide_lines = [
        "BƯỚC 1: CHUẨN BỊ ĐÁP ÁN",
        "  • Vào sheet 'Config'",
        "  • Nhập đáp án đúng cho mã đề (dòng 10, 11, ...)",
        "  • Mỗi mã đề 1 dòng, từ cột B đến BH (55 cột cho 55 câu)",
        "",
        "BƯỚC 2: NHẬP THÔNG TIN SINH VIÊN",
        "  • Vào sheet 'Chấm Điểm'",
        "  • Điền: Mã sinh viên, Họ tên, Mã đề",
        "",
        "BƯỚC 3: NHẬP ĐÁP ÁN SINH VIÊN",
        "  • Cột 'Đáp án đúng' (C): copy từ sheet Config",
        "  • Cột 'Đáp án SV' (D): dán đáp án của sinh viên",
        "  • Các công thức sẽ tự động tính điểm",
        "",
        "LƯU Ý QUAN TRỌNG:",
        "  • CLO 1 (Câu 1-40): 0.25 điểm/câu, tối đa 10 điểm",
        "  • CLO 2 (Câu 41-55): 0.5 điểm/câu, tối đa 7.5 điểm",
        "  • Tổng điểm tối đa: 17.5 điểm",
    ]
    
    for line in guide_lines:
        ws_guide.cell(row=row, column=1).value = line
        if line.startswith("BƯỚC") or line.startswith("LƯU"):
            ws_guide.cell(row=row, column=1).font = Font(bold=True, size=11)
            ws_guide.cell(row=row, column=1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        ws_guide.merge_cells(f'A{row}:B{row}')
        row += 1
    
    ws_guide.column_dimensions['A'].width = 60
    ws_guide.column_dimensions['B'].width = 20
    
    # Lưu file
    wb.save('scoring-system.xlsx')
    print("✅ File 'scoring-system.xlsx' đã tạo thành công!")
    print(f"📍 Tổng sheets: {len(wb.sheetnames)}")
    print(f"📋 Danh sách sheets: {wb.sheetnames}")

if __name__ == "__main__":
    try:
        import openpyxl
        print("🔧 Đang tạo file Excel...")
        create_scoring_system()
    except ImportError:
        print("⚠️ Cần cài đặt openpyxl:")
        print("pip install openpyxl")
        exit(1)
    except Exception as e:
        print(f"❌ Lỗi: {e}")
        import traceback
        traceback.print_exc()
